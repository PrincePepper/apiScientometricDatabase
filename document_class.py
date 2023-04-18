import time

from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from app.core import config
from app.core.config import settings


def connect_sqlalc():
    database_url = 'postgresql://{user}:{password}@{host}:{port}/{db}'.format(
        host=settings.POSTGRES_SERVER_HOST,
        port=settings.POSTGRES_SERVER_PORT,
        user=settings.POSTGRES_USERNAME,
        password=settings.POSTGRES_PASSWORD,
        db=settings.POSTGRES_DATABASE
    )
    settings.SQLALCHEMY_DATABASE_URI = database_url
    engine = create_engine(config.settings.SQLALCHEMY_DATABASE_URI, echo=True)
    return engine


engine = connect_sqlalc()
SessionLocal = sessionmaker(bind=engine)
aaa = SessionLocal()

book = load_workbook("dataset.xlsx")
sheet = book.active

# (C) IMPORT ROWS & COLUMNS
for row in range(2, sheet.max_row + 1):
    sql = "INSERT INTO users (guid, full_name, scientometric_database, document_count, citation_count, h_index, url ) " \
          f"VALUES ('{sheet.cell(row, 1).value}','{sheet.cell(row, 2).value}', '{sheet.cell(row, 3).value}','{sheet.cell(row, 4).value}','{sheet.cell(row, 5).value}','{sheet.cell(row, 6).value}','{sheet.cell(row, 7).value}')"

    aaa.execute(text(sql))
    aaa.commit()
    time.sleep(1)

aaa.close()
